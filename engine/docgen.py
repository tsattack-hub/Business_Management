"""
서류 틀 생성

docs/templates.yaml 을 읽어 hwpx(한글) 또는 xlsx 파일을 만든다.
사업 정보로 채울 수 있는 곳은 채우고, 사람이 판단해야 하는 곳은
"( 작성 필요 )" 로 비워 둔다. 빈칸을 그럴듯하게 채워 넣지 않는다.

레거시 .hwp 는 만들 수 없다. .hwpx 로 만든 뒤 한글에서 다른 이름으로 저장하면 된다.
"""
from __future__ import annotations

import datetime as dt
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

TPL_PATH = Path(__file__).resolve().parent.parent / "docs" / "templates.yaml"
FILL_MARK = "(  작성 필요  )"

FONT = "맑은 고딕"
THIN = Side(style="thin", color="AAAAAA")
BOX = Border(THIN, THIN, THIN, THIN)
HEAD_FILL = PatternFill("solid", fgColor="DCE6EF")
IN_FILL = PatternFill("solid", fgColor="FFF9DB")


@dataclass
class Template:
    id: str
    name: str
    fmt: str
    stage: str
    task: str
    citation: str
    note: str
    raw: dict


def load_templates() -> dict[str, Template]:
    doc = yaml.safe_load(TPL_PATH.read_text(encoding="utf-8")) or {}
    out: dict[str, Template] = {}
    for t in doc.get("서류", []):
        out[t["id"]] = Template(
            id=t["id"], name=t.get("명칭", t["id"]), fmt=t.get("format", "hwpx"),
            stage=t.get("단계", ""), task=t.get("태스크", ""),
            citation=t.get("근거", ""), note=t.get("설명", ""), raw=t,
        )
    return out


def templates_for(task_id: str) -> list[Template]:
    return [t for t in load_templates().values() if t.task == task_id]


# ---------------------------------------------------------------- 치환
def _won(v) -> str:
    try:
        return f"{int(round(float(v))):,}원"
    except Exception:
        return "(          원)"


def build_context(p: dict[str, Any]) -> dict[str, str]:
    """사업 정보 -> 치환자 표."""
    mat = p.get("물품분") or 0
    lab = p.get("설치분_노무") or 0
    exp = p.get("설치분_경비") or 0
    total = mat + lab + exp
    ratio = (lab + exp) / total if total else None
    return {
        "사업명": p.get("사업명") or "(사업명)",
        "연도": str(p.get("연도") or dt.date.today().year),
        "공항": p.get("공항") or "(공항)",
        "부서": p.get("부서") or "(주관부서)",
        "담당자": p.get("담당자") or "(담당자)",
        "오늘": dt.date.today().isoformat(),
        "추정가격": _won(p.get("추정가격")),
        "목표준공일": str(p.get("목표준공일") or "(준공목표일)"),
        "이행기간": str(p.get("이행기간") or "(  )"),
        "계약유형": p.get("계약유형") or "(판정 전)",
        "설치비중": f"{ratio*100:.1f}%" if ratio is not None else "(미산정)",
        "조달트랙": p.get("조달트랙") or "(미정)",
        "낙찰방법": p.get("낙찰방법") or "(미정)",
        "물품분": _won(mat) if mat else FILL_MARK,
        "설치분_노무": _won(lab) if lab else FILL_MARK,
        "설치분_경비": _won(exp) if exp else FILL_MARK,
        "설치분": _won(lab + exp) if (lab or exp) else FILL_MARK,
        "총설계금액": _won(total) if total else FILL_MARK,
    }


_PH = re.compile(r"\{([^}]+)\}")


def subst(text: Any, ctx: dict[str, str]) -> Any:
    if not isinstance(text, str):
        return text
    return _PH.sub(lambda m: ctx.get(m.group(1), m.group(0)), text)


# ---------------------------------------------------------------- hwpx
def _plan_from_blocks(tpl: Template, ctx: dict[str, str]) -> dict:
    blocks: list[dict] = []
    for b in tpl.raw.get("blocks", []):
        kind = b.get("type")

        if kind == "heading":
            blocks.append({"type": "heading", "level": int(b.get("level", 1)),
                           "text": subst(b.get("text", ""), ctx)})

        elif kind == "paragraph":
            blocks.append({"type": "paragraph", "text": subst(b.get("text", ""), ctx)})

        elif kind == "bullets":
            blocks.append({"type": "bullets",
                           "items": [subst(x, ctx) for x in b.get("items", [])]})

        elif kind == "fillin":
            blocks.append({"type": "paragraph",
                           "text": f"{FILL_MARK}  ← {subst(b.get('label',''), ctx)}"})
            if b.get("hint"):
                blocks.append({"type": "paragraph", "text": f"   ※ {subst(b['hint'], ctx)}"})

        elif kind == "table":
            cols = b.get("columns", [])
            rows = b.get("rows")
            if rows is None:
                n = int(b.get("fill", 3))
                rows = [{c["key"]: "" for c in cols} for _ in range(n)]
            else:
                rows = [{k: subst(v, ctx) for k, v in r.items()} for r in rows]
            blocks.append({"type": "table", "columns": cols, "rows": rows})

        elif kind == "page_break":
            blocks.append({"type": "page_break"})

    return {
        "schemaVersion": "hwpx.document_plan.v1",
        "title": tpl.name,
        "metadata": {
            "사업명": ctx["사업명"],
            "주관부서": ctx["부서"],
            "작성일": ctx["오늘"],
            "근거": tpl.citation or "-",
        },
        "blocks": blocks,
    }


def render_hwpx(tpl: Template, ctx: dict[str, str], out_dir: Path) -> Path:
    from hwpx_automation.office.authoring import create_document_from_plan

    plan = _plan_from_blocks(tpl, ctx)
    doc = create_document_from_plan(plan)
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / f"{tpl.name}.hwpx"
    doc.save_to_path(str(path))
    return path


# ---------------------------------------------------------------- xlsx
SETTLE_ROWS = [
    ("당초 계약금액", None, "계약서 기준", True),
    ("", None, "", False),
    ("【 사후정산 대상 】", None, "", False),
    ("산재보험료", None, "완납증명원 기준", True),
    ("고용보험료", None, "완납증명원 기준", True),
    ("국민건강보험료", None, "사업자 부담분 50%", True),
    ("국민연금보험료", None, "사업자 부담분 50%", True),
    ("노인장기요양보험료", None, "건강보험료 연동", True),
    ("퇴직공제부금비", None, "추정금액 3억원 이상인 경우", True),
    ("산업안전보건관리비", None, "사용내역 정산", True),
    ("[ 정산 감액 계 ]", "SUM", "", False),
    ("", None, "", False),
    ("【 연동 재계산 — 반드시 함께 】", None, "", False),
    ("일반관리비", "RATE", "정산 감액 × 일반관리비율", True),
    ("이윤", "RATE", "정산 감액 × 이윤율", True),
    ("[ 소     계 ]", "SUB", "", False),
    ("부가가치세", "VAT", "(정산감액 + 일반관리비 + 이윤) × 10%", False),
    ("[ 총 감액 ]", "TOTAL", "", False),
    ("", None, "", False),
    ("최종 정산 계약금액", "FINAL", "당초 계약금액 − 총 감액", False),
]


def render_xlsx(tpl: Template, ctx: dict[str, str], out_dir: Path) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    for sheet in tpl.raw.get("sheets", []):
        ws = wb.create_sheet(sheet.get("name", "Sheet1"))
        ws["A1"] = subst(sheet.get("title", tpl.name), ctx)
        ws["A1"].font = Font(name=FONT, size=15, bold=True)
        ws["A2"] = subst(sheet.get("subtitle", ""), ctx)
        ws["A2"].font = Font(name=FONT, size=10)
        ws["A3"] = f"근거 — {tpl.citation}"
        ws["A3"].font = Font(name=FONT, size=9, color="777777")

        if sheet.get("calc") == "정산":
            _build_settlement(ws)
        else:
            _build_rows(ws, sheet.get("rows", []), ctx)

    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / f"{tpl.name}.xlsx"
    wb.save(path)
    return path


def _build_rows(ws, rows: list, ctx: dict[str, str]) -> None:
    r = 5
    maxc = 1
    for row in rows:
        if not row:
            r += 1
            continue
        vals = [subst(v, ctx) for v in row]
        header = len(vals) == 1 and str(vals[0]).startswith("【")
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            cell.font = Font(name=FONT, size=10, bold=header or c == 1)
            cell.alignment = Alignment(
                horizontal="left" if c == 1 else "right", vertical="center", wrap_text=True)
            if not header:
                cell.border = BOX
            if FILL_MARK.strip() in str(v) or str(v).startswith("("):
                cell.fill = IN_FILL
        maxc = max(maxc, len(vals))
        r += 1

    for col, w in zip("ABCD", (30, 22, 12, 34)):
        ws.column_dimensions[col].width = w
    ws["A2"].alignment = Alignment(horizontal="left")


def _build_settlement(ws) -> None:
    heads = ["비        목", "금        액", "산  출  근  거"]
    for c, v in enumerate(heads, 1):
        cell = ws.cell(5, c, v)
        cell.font = Font(name=FONT, size=10, bold=True)
        cell.fill = HEAD_FILL
        cell.border = BOX
        cell.alignment = Alignment(horizontal="center", vertical="center")

    r = 6
    idx: dict[str, int] = {}
    settle_first = settle_last = None

    for label, kind, basis, editable in SETTLE_ROWS:
        if not label:
            r += 1
            continue
        ws.cell(r, 1, label).font = Font(name=FONT, size=10, bold=label.startswith(("[", "【")))
        ws.cell(r, 3, basis).font = Font(name=FONT, size=9, color="666666")
        idx[label] = r

        if kind == "SUM":
            ws.cell(r, 2, f"=SUM(B{settle_first}:B{settle_last})")
        elif kind == "RATE":
            # 사용자가 숫자를 넣어야 하는 칸. 문자열을 넣으면 아래 SUM이 깨진다.
            ws.cell(r, 2, None)
            ws.cell(r, 2).fill = IN_FILL
        elif kind == "SUB":
            ws.cell(r, 2,
                    f"=B{idx['[ 정산 감액 계 ]']}+B{idx['일반관리비']}+B{idx['이윤']}")
        elif kind == "VAT":
            ws.cell(r, 2, f"=ROUND(B{idx['[ 소     계 ]']}*0.1,0)")
        elif kind == "TOTAL":
            ws.cell(r, 2, f"=B{idx['[ 소     계 ]']}+B{idx['부가가치세']}")
        elif kind == "FINAL":
            ws.cell(r, 2, f"=B{idx['당초 계약금액']}-B{idx['[ 총 감액 ]']}")
        else:
            ws.cell(r, 2, None)
            if editable:
                ws.cell(r, 2).fill = IN_FILL
                if settle_first is None and label != "당초 계약금액":
                    settle_first = r
                if label != "당초 계약금액":
                    settle_last = r
            if label == "당초 계약금액":
                ws.cell(r, 2).fill = IN_FILL

        ws.cell(r, 2).number_format = "#,##0"
        ws.cell(r, 2).alignment = Alignment(horizontal="right")
        for c in range(1, 4):
            ws.cell(r, c).border = BOX
        r += 1

    r += 1
    for line, color in (
        ("※ 노란 셀만 입력하십시오. 나머지는 수식입니다.", "666666"),
        ("※ 보험료만 정산하고 일반관리비·이윤·부가세를 정산하지 않은 지적 사례가 있습니다.", "8E2F21"),
        ("※ 직접시공에 참여한 현장대리인은 정산 대상이 아닙니다 (간접노무비에 포함).", "8E2F21"),
        ("※ 일반관리비율·이윤율은 계약 당시 원가계산서의 요율을 그대로 적용하십시오.", "666666"),
    ):
        ws.cell(r, 1, line).font = Font(name=FONT, size=9, color=color)
        r += 1

    for col, w in zip("ABC", (30, 20, 46)):
        ws.column_dimensions[col].width = w


# ---------------------------------------------------------------- 진입점
def render(doc_id: str, project: dict[str, Any], out_dir: Path) -> Path:
    tpls = load_templates()
    if doc_id not in tpls:
        raise KeyError(f"서류 틀이 정의되지 않았습니다: {doc_id}")
    tpl = tpls[doc_id]
    ctx = build_context(project)
    if tpl.fmt == "xlsx":
        return render_xlsx(tpl, ctx, out_dir)
    return render_hwpx(tpl, ctx, out_dir)


def render_stage(stage: str, project: dict[str, Any], out_dir: Path) -> list[Path]:
    return [render(t.id, project, out_dir)
            for t in load_templates().values() if t.stage == stage]
