"""
내역서 검증 — Phase 2

AUD-030  내역서 소계·합계 검산
    R1 행 금액      금액 ≠ 수량 × 단가
    R2 합계 열      합계 ≠ 재료비 + 노무비 + 경비
    R3 소계 값      소계 ≠ 상위 명세 행의 합
    R4 SUM 범위     소계 SUM 수식이 명세 블록 전체를 덮지 않음
    R5 하드코딩     소계·합계 행인데 수식이 아님
    R6 총계 값      총계 ≠ 소계들의 합

AUD-014  수량산출서 ↔ 내역서 수량 대조
    Q1 수량 불일치
    Q2 내역서에만 있는 항목
    Q3 수량산출서에만 있는 항목

엑셀 부동소수 오차가 있으므로 1원 미만 차이는 무시한다.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from io import BytesIO

from openpyxl import load_workbook

from .boq import SUBTOTAL_WORDS, Sheet, _norm, _num
from .rules import Finding, citation_of, load_rules, rule_index

TOL = 1.0          # 1원 미만 차이는 반올림 오차로 본다
TOTAL_WORDS = ("합계", "총계", "총합계", "계")


# ---------------------------------------------------------------- 결과
@dataclass
class Issue:
    kind: str               # R1..R6 / Q1..Q3
    where: str              # 셀 주소 또는 항목명
    label: str
    expected: float | None
    actual: float | None
    message: str

    @property
    def diff(self) -> float | None:
        if self.expected is None or self.actual is None:
            return None
        return self.actual - self.expected


@dataclass
class BoqReport:
    issues: list[Issue] = field(default_factory=list)
    detail_rows: int = 0
    subtotal_rows: int = 0
    checked: int = 0
    notes: list[str] = field(default_factory=list)

    def of(self, *kinds: str) -> list[Issue]:
        return [i for i in self.issues if i.kind in kinds]


@dataclass
class QtyReport:
    issues: list[Issue] = field(default_factory=list)
    matched: int = 0
    boq_items: int = 0
    qty_items: int = 0


# ---------------------------------------------------------------- 보조
def _key(name: str, spec: str = "") -> str:
    """품명+규격을 비교용 키로. 공백·괄호·특수문자 차이를 흡수한다."""
    s = f"{name}|{spec}"
    s = re.sub(r"[\s()（）\[\]{}·,./\-_]", "", s)
    return s.lower()


def _is_word_row(label: str, words) -> bool:
    n = _norm(label)
    return bool(n) and any(n == w.replace(" ", "") for w in words)


_SUM_RE = re.compile(r"SUM\(\s*\$?([A-Z]+)\$?(\d+)\s*:\s*\$?([A-Z]+)\$?(\d+)\s*\)", re.I)


# ---------------------------------------------------------------- AUD-030
def check_boq(data: bytes, sheet_name: str, mapping: dict[str, int | None],
              header_row: int) -> BoqReport:
    """
    mapping 키: 품명 / 수량 / 재료비 / 노무비 / 경비 / 합계
                재료비단가 / 노무비단가 / 경비단가   (없으면 R1 생략)
    """
    rep = BoqReport()
    wv = load_workbook(BytesIO(data), data_only=True)
    wf = load_workbook(BytesIO(data), data_only=False)
    if sheet_name not in wv.sheetnames:
        rep.notes.append(f"시트 '{sheet_name}' 를 찾지 못했습니다.")
        return rep
    sv, sf = wv[sheet_name], wf[sheet_name]

    def col(role: str) -> int | None:
        c = mapping.get(role)
        return (c + 1) if c is not None else None      # 0-based -> openpyxl 1-based

    c_name = col("품명")
    c_qty = col("수량")
    amounts = {r: col(r) for r in ("재료비", "노무비", "경비") if col(r)}
    units = {r: col(f"{r}단가") for r in ("재료비", "노무비", "경비") if col(f"{r}단가")}
    c_sum = col("합계")

    if not amounts:
        rep.notes.append("재료비·노무비·경비 열이 지정되지 않아 검산할 수 없습니다.")
        return rep

    max_row = sv.max_row
    # ---- 행 분류
    detail: list[int] = []
    subtotals: list[int] = []
    totals: list[int] = []
    for r in range(header_row + 1, max_row + 1):
        label = _norm(sv.cell(r, c_name).value) if c_name else ""
        vals = [_num(sv.cell(r, c).value) for c in amounts.values()]
        has_amount = any(v for v in vals)
        if _is_word_row(label, ("합계", "총계", "총합계")) and has_amount:
            totals.append(r)
        elif _is_word_row(label, SUBTOTAL_WORDS) and has_amount:
            subtotals.append(r)
        elif has_amount and label and not label.startswith("【"):
            detail.append(r)

    rep.detail_rows = len(detail)
    rep.subtotal_rows = len(subtotals)

    # ---- R1 행 금액 = 수량 × 단가
    if c_qty and units:
        for r in detail:
            qty = _num(sv.cell(r, c_qty).value)
            if not qty:
                continue
            for role, ca in amounts.items():
                cu = units.get(role)
                if not cu:
                    continue
                unit = _num(sv.cell(r, cu).value)
                got = _num(sv.cell(r, ca).value)
                want = qty * unit
                rep.checked += 1
                if abs(got - want) > TOL:
                    rep.issues.append(Issue(
                        "R1", f"{sv.cell(r, ca).coordinate}",
                        f"{_norm(sv.cell(r, c_name).value)} · {role}",
                        want, got,
                        f"수량 {qty:,g} × 단가 {unit:,.0f} = {want:,.0f} 이어야 하는데 {got:,.0f} 입니다.",
                    ))

    # ---- R2 합계 열
    if c_sum:
        for r in detail:
            want = sum(_num(sv.cell(r, c).value) for c in amounts.values())
            got = _num(sv.cell(r, c_sum).value)
            rep.checked += 1
            if abs(got - want) > TOL:
                rep.issues.append(Issue(
                    "R2", sv.cell(r, c_sum).coordinate,
                    _norm(sv.cell(r, c_name).value), want, got,
                    f"재료비+노무비+경비 = {want:,.0f} 이어야 하는데 {got:,.0f} 입니다.",
                ))

    # ---- 소계별 명세 블록 확정
    blocks: list[tuple[int, list[int]]] = []
    for sr in subtotals:
        blk = [r for r in detail if r < sr and
               not any(sr2 < r for sr2 in subtotals if sr2 < sr)]
        prev = max([x for x in subtotals if x < sr], default=header_row)
        blk = [r for r in detail if prev < r < sr]
        blocks.append((sr, blk))

    check_cols = list(amounts.values()) + ([c_sum] if c_sum else [])

    for sr, blk in blocks:
        if not blk:
            continue
        for c in check_cols:
            cell_v = sv.cell(sr, c)
            cell_f = sf.cell(sr, c)
            want = sum(_num(sv.cell(r, c).value) for r in blk)
            got = _num(cell_v.value)
            rep.checked += 1

            # R5 하드코딩
            f = cell_f.value
            if not (isinstance(f, str) and f.startswith("=")):
                rep.issues.append(Issue(
                    "R5", cell_v.coordinate, f"소계 ({sv.cell(sr, c_name).value if c_name else ''})",
                    want, got,
                    "수식이 아니라 값이 입력되어 있습니다. 원본이 바뀌어도 갱신되지 않습니다.",
                ))
            else:
                # R4 SUM 범위 누락
                m = _SUM_RE.search(f)
                if m:
                    r1, r2 = int(m.group(2)), int(m.group(4))
                    if r1 > min(blk) or r2 < max(blk):
                        missing = [r for r in blk if not (r1 <= r <= r2)]
                        rep.issues.append(Issue(
                            "R4", cell_v.coordinate,
                            f"소계 SUM 범위", None, None,
                            f"수식 범위 {r1}~{r2} 행이 명세 블록 {min(blk)}~{max(blk)} 행을 덮지 못합니다. "
                            f"누락된 행: {', '.join(map(str, missing))}",
                        ))

            # R3 소계 값
            if abs(got - want) > TOL:
                rep.issues.append(Issue(
                    "R3", cell_v.coordinate, "소계", want, got,
                    f"명세 {len(blk)}개 행의 합은 {want:,.0f} 인데 소계는 {got:,.0f} 입니다.",
                ))

    # ---- R6 총계
    for tr in totals:
        for c in check_cols:
            want = sum(_num(sv.cell(sr, c).value) for sr, _ in blocks)
            got = _num(sv.cell(tr, c).value)
            rep.checked += 1
            f = sf.cell(tr, c).value
            if not (isinstance(f, str) and f.startswith("=")):
                rep.issues.append(Issue(
                    "R5", sv.cell(tr, c).coordinate, "총계",
                    want, got,
                    "총계가 수식이 아니라 값입니다. 소계가 바뀌어도 따라오지 않습니다.",
                ))
            if blocks and abs(got - want) > TOL:
                rep.issues.append(Issue(
                    "R6", sv.cell(tr, c).coordinate, "총계", want, got,
                    f"소계 합은 {want:,.0f} 인데 총계는 {got:,.0f} 입니다. "
                    f"차액 {got - want:+,.0f}원.",
                ))
    wv.close()
    wf.close()
    return rep


# ---------------------------------------------------------------- AUD-014
def check_quantity(boq_sheet: Sheet, boq_map: dict[str, int | None],
                   qty_sheet: Sheet, qty_map: dict[str, int | None]) -> QtyReport:
    """수량산출서와 내역서의 수량을 품명+규격 키로 대조."""
    rep = QtyReport()

    def collect(sh: Sheet, m: dict[str, int | None]) -> dict[str, tuple[str, float]]:
        out: dict[str, tuple[str, float]] = {}
        cn, cs, cq = m.get("품명"), m.get("규격"), m.get("수량")
        if cn is None or cq is None:
            return out
        for row in sh.rows:
            if cn >= len(row) or cq >= len(row):
                continue
            name = _norm(row[cn])
            if not name or name.startswith("【") or _is_word_row(name, SUBTOTAL_WORDS + TOTAL_WORDS):
                continue
            q = _num(row[cq])
            if q == 0:
                continue
            spec = _norm(row[cs]) if cs is not None and cs < len(row) else ""
            out[_key(name, spec)] = (f"{name} ({spec})" if spec else name, q)
        return out

    a = collect(boq_sheet, boq_map)      # 내역서
    b = collect(qty_sheet, qty_map)      # 수량산출서
    rep.boq_items, rep.qty_items = len(a), len(b)

    for k, (label, qa) in a.items():
        if k not in b:
            rep.issues.append(Issue(
                "Q2", label, label, None, qa,
                f"내역서에는 있으나 수량산출서에 없습니다 (내역서 수량 {qa:,g}).",
            ))
            continue
        qb = b[k][1]
        if abs(qa - qb) > 1e-6:
            rep.issues.append(Issue(
                "Q1", label, label, qb, qa,
                f"내역서 {qa:,g} vs 수량산출서 {qb:,g} — 차이 {qa - qb:+,g}.",
            ))
        else:
            rep.matched += 1

    for k, (label, qb) in b.items():
        if k not in a:
            rep.issues.append(Issue(
                "Q3", label, label, qb, None,
                f"수량산출서에는 있으나 내역서에 없습니다 (산출 수량 {qb:,g}).",
            ))
    return rep


# ---------------------------------------------------------------- Finding 변환
def to_findings(boq_rep: BoqReport | None, qty_rep: QtyReport | None) -> list[Finding]:
    idx = rule_index(load_rules())
    out: list[Finding] = []

    if boq_rep is not None:
        r = idx.get("AUD-030", {})
        n = len(boq_rep.issues)
        out.append(Finding(
            rule_id="AUD-030", name=r.get("명칭", "내역서 소계·합계 검산"),
            severity=r.get("심각도", "error"), passed=(n == 0),
            message=(f"검산 {boq_rep.checked}건 중 {n}건 불일치."
                     if n else f"검산 {boq_rep.checked}건 모두 일치합니다."),
            citation=citation_of(r), status=r.get("status", "confirmed"),
            detail=[f"{i.kind} · {i.where} · {i.message}" for i in boq_rep.issues[:40]],
        ))

    if qty_rep is not None:
        r = idx.get("AUD-014", {})
        n = len(qty_rep.issues)
        out.append(Finding(
            rule_id="AUD-014", name=r.get("명칭", "수량산출서와 내역서의 수량 불일치"),
            severity=r.get("심각도", "error"), passed=(n == 0),
            message=(f"{qty_rep.boq_items}개 항목 대조 — 일치 {qty_rep.matched}, 문제 {n}건."
                     if n else f"{qty_rep.matched}개 항목 수량이 모두 일치합니다."),
            citation=citation_of(r), status=r.get("status", "confirmed"),
            detail=[f"{i.kind} · {i.label} · {i.message}" for i in qty_rep.issues[:40]],
        ))
    return out
