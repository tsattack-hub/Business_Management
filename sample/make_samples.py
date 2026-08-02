"""
샘플 내역서 생성기

두 벌을 만든다.
  sample_내역서_정상.xlsx   — 오류 없음. 검증기가 통과시켜야 한다.
  sample_내역서_오류.xlsx   — 실제로 자주 나오는 오류 6종을 심었다.

심어 둔 오류 (검증기가 전부 잡아야 함)
  E1  소계 SUM 범위 누락      — 마지막 명세 행이 합계에서 빠짐
  E2  하드코딩된 총계          — 수식이 아니라 숫자를 박아 넣음. 실제 합과 불일치
  E3  행 금액 오류             — 금액 ≠ 수량 × 단가
  E4  합계 열 오류             — 합계 ≠ 재료비 + 노무비 + 경비
  E5  수량 불일치 2건          — 수량산출서와 내역서의 수량이 다름
  E6  누락 항목                — 한쪽에만 있는 항목 (양방향 각 1건)

구조는 공공 발주 내역서의 통상 형태를 따랐다.
  원가계산서 / 내역서(2단 병합 헤더) / 수량산출서 / 일위대가표
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent
FONT = "Arial"

THIN = Side(style="thin", color="AAAAAA")
MED = Side(style="medium", color="555555")
BOX = Border(THIN, THIN, THIN, THIN)
HEAD_FILL = PatternFill("solid", fgColor="DCE6EF")
SUB_FILL = PatternFill("solid", fgColor="F2F5F8")
IN_FILL = PatternFill("solid", fgColor="FFF9DB")

# ---------------------------------------------------------------- 원자료
# (공종, 품명, 규격, 단위, 수량, 재료비단가, 노무비단가, 경비단가)
ITEMS = [
    ("영상감시설비", "네트워크 카메라(고정형)", "200만화소 IP, 실외형 IP66", "대", 18, 1_250_000, 180_000, 0),
    ("영상감시설비", "네트워크 카메라(PTZ)", "200만화소 IP, 32배줌", "대", 6, 3_150_000, 240_000, 0),
    ("영상감시설비", "영상저장장치", "32ch NVR, RAID5", "대", 2, 8_400_000, 420_000, 0),
    ("영상감시설비", "저장용 하드디스크", "8TB SATA, 감시용", "개", 12, 310_000, 35_000, 0),
    ("영상감시설비", "통합관제 소프트웨어", "VMS 64ch 라이선스", "식", 1, 12_000_000, 1_800_000, 0),
    ("영상감시설비", "카메라 브래킷", "벽부형 알루미늄", "개", 18, 45_000, 32_000, 0),
    ("영상감시설비", "카메라 폴", "STK 3.0m, 용융아연도금", "개", 6, 420_000, 165_000, 0),
    ("배관배선공사", "후강전선관", "22C, 아연도강관", "m", 1800, 3_200, 4_100, 0),
    ("배관배선공사", "케이블 트레이", "와이어메쉬 100W", "m", 320, 11_000, 9_500, 0),
    ("배관배선공사", "UTP 케이블", "Cat.6 UTP 4P", "m", 2400, 900, 1_100, 0),
    ("배관배선공사", "광케이블", "SM 12C, 옥외용", "m", 600, 2_400, 1_800, 0),
    ("배관배선공사", "접지공사", "제3종 접지 100Ω 이하", "개소", 12, 58_000, 86_000, 0),
    ("부대공사", "고소작업차 사용", "28m 스카이차", "일", 14, 0, 0, 420_000),
    ("부대공사", "시험 및 조정", "계통 시험, 화각 조정", "식", 1, 0, 2_400_000, 650_000),
    ("부대공사", "폐기물 처리", "기존 설비 철거분", "식", 1, 0, 0, 380_000),
]

# 수량산출서 산출근거
BASIS = {
    "네트워크 카메라(고정형)": "여객터미널 1층 8 + 2층 6 + 주차장 4 = 18대",
    "네트워크 카메라(PTZ)": "출입구 4 + 계류장측 2 = 6대",
    "영상저장장치": "36ch ÷ 32ch = 2대 (예비 채널 포함)",
    "저장용 하드디스크": "NVR 2대 × 6bay = 12개",
    "통합관제 소프트웨어": "64ch 라이선스 1식",
    "카메라 브래킷": "고정형 카메라 18대 × 1 = 18개",
    "카메라 폴": "PTZ 카메라 6대 × 1 = 6개",
    "후강전선관": "1층 720 + 2층 640 + 옥외 440 = 1,800m",
    "케이블 트레이": "주배선 간선 320m",
    "UTP 케이블": "카메라 24대 × 평균 100m = 2,400m",
    "광케이블": "관제실 ↔ 층별 집선함 3구간 × 200m = 600m",
    "접지공사": "집선함 6 + NVR 2 + 폴 4 = 12개소",
    "고소작업차 사용": "옥외 설치 10일 + 시험 4일 = 14일",
    "시험 및 조정": "전 계통 1식",
    "폐기물 처리": "기존 아날로그 카메라 철거분 1식",
}


# ---------------------------------------------------------------- 유틸
def style_range(ws, r1, r2, c1, c2, *, size=10, bold=False, fill=None):
    for row in ws.iter_rows(min_row=r1, max_row=r2, min_col=c1, max_col=c2):
        for c in row:
            c.font = Font(name=FONT, size=size, bold=bold)
            c.border = BOX
            if fill:
                c.fill = fill


def widths(ws, spec: dict[str, int]):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


# ---------------------------------------------------------------- 내역서
def build_boq(wb, errors: bool):
    ws = wb.create_sheet("내역서")
    ws["A1"] = "OO공항 여객터미널 CCTV 구매설치"
    ws["A1"].font = Font(name=FONT, size=14, bold=True)
    ws["A2"] = "공 종 별 내 역 서"
    ws["A2"].font = Font(name=FONT, size=11)
    ws["A3"] = "(단위: 원, VAT 별도)"
    ws["A3"].font = Font(name=FONT, size=9, color="777777")

    top = ["품    명", "규    격", "단위", "수량", "재 료 비", "", "노 무 비", "", "경    비", "", "합    계"]
    sub = ["", "", "", "", "단가", "금액", "단가", "금액", "단가", "금액", "금액"]
    for c, v in enumerate(top, 1):
        ws.cell(5, c, v)
    for c, v in enumerate(sub, 1):
        ws.cell(6, c, v)
    for col in ("A", "B", "C", "D", "K"):
        ws.merge_cells(f"{col}5:{col}6")
    for a, b in (("E", "F"), ("G", "H"), ("I", "J")):
        ws.merge_cells(f"{a}5:{b}5")
    style_range(ws, 5, 6, 1, 11, bold=True, fill=HEAD_FILL)
    for row in ws.iter_rows(min_row=5, max_row=6, max_col=11):
        for c in row:
            c.alignment = Alignment(horizontal="center", vertical="center")

    r = 7
    subtotal_rows: list[tuple[str, int, int, int]] = []   # (공종, 시작, 끝, 소계행)
    detail_rows: list[int] = []
    trades = []
    for it in ITEMS:
        if it[0] not in trades:
            trades.append(it[0])

    for trade in trades:
        ws.cell(r, 1, f"【 {trade} 】").font = Font(name=FONT, size=10, bold=True)
        ws.cell(r, 1).fill = SUB_FILL
        style_range(ws, r, r, 1, 11, fill=SUB_FILL)
        r += 1
        first = r
        for _t, nm, spec, unit, qty, mu, lu, eu in [i for i in ITEMS if i[0] == trade]:
            ws.cell(r, 1, nm)
            ws.cell(r, 2, spec)
            ws.cell(r, 3, unit)
            ws.cell(r, 4, qty)
            ws.cell(r, 5, mu)
            ws.cell(r, 6, f"=D{r}*E{r}")
            ws.cell(r, 7, lu)
            ws.cell(r, 8, f"=D{r}*G{r}")
            ws.cell(r, 9, eu)
            ws.cell(r, 10, f"=D{r}*I{r}")
            ws.cell(r, 11, f"=F{r}+H{r}+J{r}")
            detail_rows.append(r)
            r += 1
        last = r - 1

        # ---- E1: 배관배선공사 소계의 SUM 범위에서 마지막 행 누락
        end = last - 1 if (errors and trade == "배관배선공사") else last
        ws.cell(r, 1, "소      계").font = Font(name=FONT, size=10, bold=True)
        for col in (6, 8, 10, 11):
            L = get_column_letter(col)
            ws.cell(r, col, f"=SUM({L}{first}:{L}{end})")
        style_range(ws, r, r, 1, 11, bold=True, fill=SUB_FILL)
        subtotal_rows.append((trade, first, last, r))
        r += 2

    # ---- 총계
    total_row = r
    ws.cell(r, 1, "합      계").font = Font(name=FONT, size=11, bold=True)
    if errors:
        # ---- E2: 총계를 수식이 아닌 숫자로 박아 넣음 (실제 합과 다름)
        ws.cell(r, 6, 138_500_000)
        ws.cell(r, 8, 41_200_000)
        ws.cell(r, 10, 7_500_000)
        ws.cell(r, 11, 187_200_000)
    else:
        for col in (6, 8, 10, 11):
            L = get_column_letter(col)
            parts = "+".join(f"{L}{s[3]}" for s in subtotal_rows)
            ws.cell(r, col, f"={parts}")
    style_range(ws, r, r, 1, 11, size=11, bold=True, fill=HEAD_FILL)
    for c in range(1, 12):
        ws.cell(r, c).border = Border(MED, MED, MED, MED)

    # ---- E3 / E4: 행 단위 오류
    if errors:
        for row in detail_rows:
            nm = ws.cell(row, 1).value
            if nm == "UTP 케이블":
                ws.cell(row, 6, 2_400_000)          # E3: 2400×900=2,160,000 이어야 함
            if nm == "시험 및 조정":
                ws.cell(row, 11, f"=F{row}+H{row}")  # E4: 경비 누락 (경비 650,000 빠짐)

    for row in ws.iter_rows(min_row=7, max_row=total_row, max_col=11):
        for c in row:
            c.border = BOX
            if c.font.name != FONT:
                c.font = Font(name=FONT, size=10)
            if c.column in (3, 4):
                c.alignment = Alignment(horizontal="center")
            elif c.column <= 2:
                c.alignment = Alignment(horizontal="left")
            else:
                c.alignment = Alignment(horizontal="right")
            if c.column >= 4:
                c.number_format = "#,##0"

    widths(ws, {"A": 26, "B": 30, "C": 7, "D": 10, "E": 13, "F": 15,
                "G": 12, "H": 14, "I": 12, "J": 14, "K": 16})
    ws.freeze_panes = "A7"
    return subtotal_rows, total_row


# ---------------------------------------------------------------- 수량산출서
def build_qty(wb, errors: bool):
    ws = wb.create_sheet("수량산출서")
    ws["A1"] = "수 량 산 출 서"
    ws["A1"].font = Font(name=FONT, size=14, bold=True)
    ws["A2"] = "※ 산출 근거를 명확히 기재할 것. 순서는 내역서와 일치시킬 것."
    ws["A2"].font = Font(name=FONT, size=9, color="777777")

    heads = ["연번", "품    명", "규    격", "단위", "산  출  근  거", "수량"]
    for c, v in enumerate(heads, 1):
        ws.cell(4, c, v)
    style_range(ws, 4, 4, 1, 6, bold=True, fill=HEAD_FILL)
    for c in range(1, 7):
        ws.cell(4, c).alignment = Alignment(horizontal="center", vertical="center")

    r, n = 5, 0
    for _t, nm, spec, unit, qty, *_ in ITEMS:
        # ---- E6: 내역서에만 있고 산출서에 없는 항목
        if errors and nm == "폐기물 처리":
            continue
        q = qty
        # ---- E5: 수량 불일치 2건
        if errors and nm == "후강전선관":
            q = 1650
        if errors and nm == "접지공사":
            q = 14
        n += 1
        ws.cell(r, 1, n)
        ws.cell(r, 2, nm)
        ws.cell(r, 3, spec)
        ws.cell(r, 4, unit)
        ws.cell(r, 5, BASIS.get(nm, ""))
        ws.cell(r, 6, q)
        r += 1

    # ---- E6: 산출서에만 있고 내역서에 없는 항목
    if errors:
        n += 1
        ws.cell(r, 1, n)
        ws.cell(r, 2, "방수형 접속함")
        ws.cell(r, 3, "IP66, 200×300")
        ws.cell(r, 4, "개")
        ws.cell(r, 5, "층별 집선 6개소 × 1 = 6개")
        ws.cell(r, 6, 6)
        r += 1

    style_range(ws, 5, r - 1, 1, 6)
    for row in ws.iter_rows(min_row=5, max_row=r - 1, max_col=6):
        for c in row:
            c.alignment = Alignment(
                horizontal="center" if c.column in (1, 4, 6) else "left", vertical="center")
            if c.column == 6:
                c.number_format = "#,##0"
    widths(ws, {"A": 6, "B": 26, "C": 30, "D": 7, "E": 46, "F": 10})
    ws.freeze_panes = "A5"


# ---------------------------------------------------------------- 일위대가
def build_unit(wb):
    ws = wb.create_sheet("일위대가표")
    ws["A1"] = "일 위 대 가 표"
    ws["A1"].font = Font(name=FONT, size=14, bold=True)
    ws["A2"] = "※ 품셈에 없는 항목은 작업시간 기준 공량 산정 근거를 별도 첨부"
    ws["A2"].font = Font(name=FONT, size=9, color="777777")

    heads = ["호별", "품    명", "규    격", "단위", "수량", "단가", "금액", "비고"]
    for c, v in enumerate(heads, 1):
        ws.cell(4, c, v)
    style_range(ws, 4, 4, 1, 8, bold=True, fill=HEAD_FILL)
    for c in range(1, 9):
        ws.cell(4, c).alignment = Alignment(horizontal="center", vertical="center")

    blocks = [
        ("제1호", "네트워크 카메라 설치 (고정형)", "대당", [
            ("정보통신기사", "인", 0.45, 285_000, "정보통신 표준품셈"),
            ("정보통신기능사", "인", 0.30, 175_000, "정보통신 표준품셈"),
            ("공구손료", "식", 1.0, 4_200, "인력품의 3%"),
        ]),
        ("제2호", "후강전선관 배관 (22C)", "m당", [
            ("정보통신기능사", "인", 0.018, 175_000, "노출 배관 기준"),
            ("보통인부", "인", 0.009, 158_000, ""),
            ("잡재료비", "식", 1.0, 96, "재료비의 3%"),
        ]),
        ("제3호", "UTP 케이블 포설 (Cat.6)", "m당", [
            ("정보통신기능사", "인", 0.005, 175_000, "관내 포설"),
            ("보통인부", "인", 0.0014, 158_000, ""),
        ]),
    ]

    r = 5
    for code, title, unit, rows in blocks:
        ws.cell(r, 1, code).font = Font(name=FONT, size=10, bold=True)
        ws.cell(r, 2, title).font = Font(name=FONT, size=10, bold=True)
        ws.cell(r, 4, unit)
        style_range(ws, r, r, 1, 8, bold=True, fill=SUB_FILL)
        r += 1
        first = r
        for nm, u, qty, price, note in rows:
            ws.cell(r, 2, nm)
            ws.cell(r, 4, u)
            ws.cell(r, 5, qty)
            ws.cell(r, 6, price)
            ws.cell(r, 7, f"=ROUND(E{r}*F{r},0)")
            ws.cell(r, 8, note)
            r += 1
        ws.cell(r, 2, "계").font = Font(name=FONT, size=10, bold=True)
        ws.cell(r, 7, f"=SUM(G{first}:G{r-1})")
        style_range(ws, r, r, 1, 8, bold=True)
        r += 2

    style_range(ws, 5, r - 1, 1, 8)
    for row in ws.iter_rows(min_row=5, max_row=r - 1, max_col=8):
        for c in row:
            if c.column in (5, 6, 7):
                c.number_format = "#,##0.###" if c.column == 5 else "#,##0"
                c.alignment = Alignment(horizontal="right")
            elif c.column in (1, 4):
                c.alignment = Alignment(horizontal="center")
    widths(ws, {"A": 8, "B": 30, "C": 22, "D": 8, "E": 10, "F": 12, "G": 14, "H": 22})


# ---------------------------------------------------------------- 원가계산서
def build_cost(wb, total_row: int):
    ws = wb.create_sheet("원가계산서", 0)
    ws["A1"] = "구 매 설 치 원 가 계 산 서"
    ws["A1"].font = Font(name=FONT, size=15, bold=True)
    ws["A2"] = "사업명: OO공항 여객터미널 CCTV 구매설치"
    ws["A2"].font = Font(name=FONT, size=10)
    ws["A3"] = "※ 노란 셀만 입력. 나머지는 수식입니다."
    ws["A3"].font = Font(name=FONT, size=9, color="777777")

    for c, v in enumerate(["비        목", "금        액", "산  출  근  거"], 1):
        ws.cell(5, c, v)
    style_range(ws, 5, 5, 1, 3, bold=True, fill=HEAD_FILL)
    for c in range(1, 4):
        ws.cell(5, c).alignment = Alignment(horizontal="center", vertical="center")

    B = f"내역서!F{total_row}"
    L = f"내역서!H{total_row}"
    E = f"내역서!J{total_row}"

    # 행 번호를 하드코딩하지 않는다. 비목 이름 -> 행 번호를 먼저 잡고 수식을 만든다.
    names = [
        "직접재료비", "직접노무비", "간접노무비", "[ 노무비 계 ]", "기계경비",
        "산재보험료", "고용보험료", "국민건강보험료", "국민연금보험료", "노인장기요양보험료",
        "산업안전보건관리비", "기타경비", "[ 경비 계 ]", "[ 순 원 가 ]",
        "일반관리비", "이윤", "[ 총 원 가 ]", "부가가치세", "[ 도 급 액 ]",
    ]
    R = {n: 6 + i for i, n in enumerate(names)}

    def b(n: str) -> str:
        return f"B{R[n]}"

    exp_first, exp_last = R["기계경비"], R["기타경비"]

    rows = [
        ("직접재료비", f"={B}", "내역서 재료비 합계"),
        ("직접노무비", f"={L}", "내역서 노무비 합계"),
        ("간접노무비", f"=ROUND({b('직접노무비')}*0.101,0)", "직접노무비 × 10.1%  (설치분에만 적용)"),
        ("[ 노무비 계 ]", f"={b('직접노무비')}+{b('간접노무비')}", ""),
        ("기계경비", f"={E}", "내역서 경비 합계"),
        ("산재보험료", f"=ROUND({b('[ 노무비 계 ]')}*0.0389,0)", "노무비 계 × 3.89%"),
        ("고용보험료", f"=ROUND({b('[ 노무비 계 ]')}*0.0087,0)", "노무비 계 × 0.87%"),
        ("국민건강보험료", f"=ROUND({b('직접노무비')}*0.0397,0)", "직접노무비 × 3.97%"),
        ("국민연금보험료", f"=ROUND({b('직접노무비')}*0.045,0)", "직접노무비 × 4.5%"),
        ("노인장기요양보험료", f"=ROUND({b('국민건강보험료')}*0.1295,0)", "건강보험료 × 12.95%"),
        ("산업안전보건관리비",
         f"=ROUND(({b('직접재료비')}+{b('직접노무비')})*0.0187,0)", "(재료비+직접노무비) × 1.87%"),
        ("기타경비",
         f"=ROUND(({b('직접재료비')}+{b('[ 노무비 계 ]')})*0.058,0)", "(재료비+노무비계) × 5.8%"),
        ("[ 경비 계 ]", f"=SUM(B{exp_first}:B{exp_last})", ""),
        ("[ 순 원 가 ]",
         f"={b('직접재료비')}+{b('[ 노무비 계 ]')}+{b('[ 경비 계 ]')}", ""),
        ("일반관리비", f"=ROUND({b('[ 순 원 가 ]')}*0.05,0)", "순원가 × 5%"),
        ("이윤",
         f"=ROUND(({b('[ 노무비 계 ]')}+{b('[ 경비 계 ]')}+{b('일반관리비')})*0.09,0)",
         "(노무비+경비+일반관리비) × 9%"),
        ("[ 총 원 가 ]",
         f"={b('[ 순 원 가 ]')}+{b('일반관리비')}+{b('이윤')}", ""),
        ("부가가치세", f"=ROUND({b('[ 총 원 가 ]')}*0.1,0)", "총원가 × 10%"),
        ("[ 도 급 액 ]", f"={b('[ 총 원 가 ]')}+{b('부가가치세')}", ""),
    ]
    assert [x[0] for x in rows] == names, "비목 순서가 R 인덱스와 어긋났습니다"
    r = 6
    for name, formula, basis in rows:
        ws.cell(r, 1, name)
        ws.cell(r, 2, formula)
        ws.cell(r, 3, basis)
        bold = name.startswith("[")
        ws.cell(r, 1).font = Font(name=FONT, size=10, bold=bold)
        ws.cell(r, 2).font = Font(name=FONT, size=10, bold=bold)
        ws.cell(r, 3).font = Font(name=FONT, size=9, color="666666")
        ws.cell(r, 2).number_format = "#,##0"
        ws.cell(r, 2).alignment = Alignment(horizontal="right")
        for c in range(1, 4):
            ws.cell(r, c).border = BOX
            if bold:
                ws.cell(r, c).fill = SUB_FILL
        r += 1

    ws.cell(r + 1, 1, "적용 제비율").font = Font(name=FONT, size=10, bold=True)
    ws.cell(r + 2, 1, "조달청 원가계산 제비율 (공사규모 3억 미만 / 공사기간 7~12월) 기준")
    ws.cell(r + 2, 1).font = Font(name=FONT, size=9, color="666666")
    ws.cell(r + 3, 1, "※ 값은 예시입니다. 실제 적용 시 조달청 최신본으로 대조하십시오. (AUD-015)")
    ws.cell(r + 3, 1).font = Font(name=FONT, size=9, color="8E2F21")

    widths(ws, {"A": 26, "B": 18, "C": 46})


# ---------------------------------------------------------------- 조립
def make(errors: bool, filename: str):
    wb = Workbook()
    wb.remove(wb.active)
    subtotals, total_row = build_boq(wb, errors)
    build_qty(wb, errors)
    build_unit(wb)
    build_cost(wb, total_row)
    path = OUT / filename
    wb.save(path)
    return path, subtotals, total_row


if __name__ == "__main__":
    for err, fn in ((False, "sample_내역서_정상.xlsx"), (True, "sample_내역서_오류.xlsx")):
        p, subs, tr = make(err, fn)
        print(f"{p.name}  소계 {len(subs)}개, 총계 {tr}행")
