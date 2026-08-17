"""
내역서(xlsx) 파싱 · 물품분 / 설치분 집계

현실 인식
  발주부서마다 내역서 서식이 다르다. 자동 인식은 '거들 뿐'이고,
  최종 결정은 사용자의 열 매핑이다. 인식 실패 시 직접입력 경로가 항상 열려 있어야 한다.

MX-001 산정 기준
  물품분 = Σ 재료비
  설치분 = Σ 노무비 + Σ 경비
"""
from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO

from openpyxl import load_workbook

# 열 이름 자동 인식 키워드 (앞쪽일수록 우선)
HINTS: dict[str, list[str]] = {
    "품명": ["품명", "공종", "품 명", "명칭", "규격및명칭", "공사명"],
    "규격": ["규격", "규 격", "형식", "사양"],
    "단위": ["단위", "단 위"],
    "수량": ["수량", "수 량", "물량"],
    "재료비": ["재료비", "재 료 비", "자재비", "재료"],
    "노무비": ["노무비", "노 무 비", "인건비", "노무"],
    "경비": ["경비", "경 비", "기계경비", "간접비"],
    "합계": ["합계", "합 계", "계", "금액", "소계"],
}

# 금액 열은 보통 (단가, 금액) 쌍으로 나온다. '금액' 쪽을 잡아야 한다.
AMOUNT_HINT = ["금액", "금 액", "계", "합계"]
UNITPRICE_HINT = ["단가", "단 가"]


@dataclass
class Column:
    index: int          # 0-based
    letter: str
    header: str
    guess: str = ""     # 추론된 역할


@dataclass
class Sheet:
    name: str
    header_row: int             # 1-based
    columns: list[Column]
    rows: list[list]            # 헤더 아래 데이터 원본
    n_rows: int = 0


@dataclass
class Totals:
    material: float = 0.0       # 물품분 (재료비)
    labor: float = 0.0          # 설치 노무비
    expense: float = 0.0        # 설치 경비
    line_count: int = 0
    warnings: list[str] = field(default_factory=list)

    @property
    def install(self) -> float:
        return self.labor + self.expense

    @property
    def total(self) -> float:
        return self.material + self.install

    @property
    def install_ratio(self) -> float | None:
        return (self.install / self.total) if self.total else None


def _col_letter(idx: int) -> str:
    s, n = "", idx + 1
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def _norm(v) -> str:
    return str(v).replace(" ", "").replace("\n", "").strip() if v is not None else ""


def _ffill(values: list, span: int = 3) -> list:
    """
    병합된 헤더 대응. '재료비'가 E:F에 병합돼 있으면 F는 None으로 읽히므로,
    직전 값을 최대 span칸까지 흘려보낸다.
    span을 두는 이유 — 무제한으로 채우면 무관한 열까지 오염된다.
    """
    out, last, run = [], None, 0
    for v in values:
        s = _norm(v)
        if s:
            last, run = s, 0
            out.append(s)
        elif last and run < span:
            run += 1
            out.append(last)
        else:
            out.append("")
    return out


def _score_header_row(values: list) -> int:
    """헤더 후보 행 점수: 인식 키워드가 몇 개나 걸리는가."""
    joined = [_norm(v) for v in values]
    score = 0
    for hints in HINTS.values():
        if any(any(h.replace(" ", "") in cell for cell in joined if cell) for h in hints):
            score += 1
    return score


def read_workbook(data: bytes, max_scan: int = 25) -> list[Sheet]:
    """업로드된 xlsx에서 시트별 헤더 행과 열을 추정한다."""
    wb = load_workbook(BytesIO(data), data_only=True, read_only=True)
    sheets: list[Sheet] = []
    for ws in wb.worksheets:
        grid = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            grid.append(list(row))
            if i > 4000:
                break
        if not grid:
            continue

        best_row, best_score = 0, -1
        for i in range(min(max_scan, len(grid))):
            s = _score_header_row(grid[i])
            if s > best_score:
                best_row, best_score = i, s
        if best_score <= 0:
            best_row = 0

        # 병합 헤더 대응: 상위 행은 전방 채움, 하위 행은 그대로 붙인다
        width = max(len(grid[r]) for r in range(best_row, min(best_row + 2, len(grid))))
        header = _ffill(list(grid[best_row]) + [None] * (width - len(grid[best_row])))
        above = _ffill(list(grid[best_row - 1])) if best_row > 0 else []
        below = grid[best_row + 1] if best_row + 1 < len(grid) else []

        columns: list[Column] = []
        for idx in range(width):
            parts = []
            if idx < len(above) and above[idx]:
                parts.append(above[idx])
            if idx < len(header) and header[idx]:
                parts.append(header[idx])
            if idx < len(below) and below[idx] is not None:
                parts.append(_norm(below[idx]))
            # 전방 채움으로 같은 말이 반복되면 하나만 남긴다
            seen, uniq = set(), []
            for p in parts:
                if p and p not in seen:
                    seen.add(p)
                    uniq.append(p)
            label = " ".join(uniq) or f"({_col_letter(idx)}열)"
            columns.append(Column(index=idx, letter=_col_letter(idx), header=label))

        _guess_roles(columns)
        data_rows = grid[best_row + 1:]
        sheets.append(Sheet(
            name=ws.title, header_row=best_row + 1, columns=columns,
            rows=data_rows, n_rows=len(data_rows),
        ))
    wb.close()
    return sheets


def _guess_roles(columns: list[Column]) -> None:
    """
    열 역할 추론.
    금액 열과 단가 열을 함께 잡는다 — 행 단위 검산(금액 = 수량 × 단가)에 둘 다 필요하다.
    """
    taken: set[int] = set()

    for role in ("재료비", "노무비", "경비"):
        amt_c: Column | None = None
        unit_c: Column | None = None
        for c in columns:
            if c.index in taken:
                continue
            h = c.header.replace(" ", "")
            if not any(hint.replace(" ", "") in h for hint in HINTS[role]):
                continue
            is_unit = any(u.replace(" ", "") in h for u in UNITPRICE_HINT)
            is_amt = any(a.replace(" ", "") in h for a in AMOUNT_HINT)
            if is_unit and unit_c is None:
                unit_c = c
            elif is_amt and amt_c is None:
                amt_c = c
            elif amt_c is None and not is_unit:
                amt_c = c
        # 금액 열을 못 찾았고 단가 열만 있으면, 그 열을 금액으로 쓰지 않는다.
        # (단가를 금액으로 오인하면 설치비중이 통째로 틀린다)
        if amt_c is not None:
            amt_c.guess = role
            taken.add(amt_c.index)
        if unit_c is not None:
            unit_c.guess = f"{role}단가"
            taken.add(unit_c.index)

    for role in ("품명", "규격", "단위", "수량", "합계"):
        for c in columns:
            if c.guess or c.index in taken:
                continue
            h = c.header.replace(" ", "")
            if any(hint.replace(" ", "") in h for hint in HINTS[role]):
                c.guess = role
                taken.add(c.index)
                break


def _num(v) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").replace("₩", "").replace("원", "").strip()
    if not s or s in ("-", "—"):
        return 0.0
    neg = s.startswith("(") and s.endswith(")")
    if neg:
        s = s[1:-1]
    try:
        val = float(s)
    except ValueError:
        return 0.0
    return -val if neg else val


SUBTOTAL_WORDS = ("소계", "합계", "총계", "계", "누계", "총합계", "원가계산", "합 계", "소 계")


def aggregate(sheet: Sheet, mapping: dict[str, int | None],
              skip_subtotals: bool = True) -> Totals:
    """
    mapping: {'재료비': 열index, '노무비': 열index, '경비': 열index, '품명': 열index|None}
    소계·합계 행은 이중계상을 막기 위해 기본적으로 제외한다.
    """
    t = Totals()
    name_col = mapping.get("품명")
    skipped = 0

    for row in sheet.rows:
        if not any(c is not None and str(c).strip() != "" for c in row):
            continue
        label = ""
        if name_col is not None and name_col < len(row):
            label = _norm(row[name_col])

        if skip_subtotals and label and any(w.replace(" ", "") == label for w in SUBTOTAL_WORDS):
            skipped += 1
            continue

        m = _num(row[mapping["재료비"]]) if mapping.get("재료비") is not None and mapping["재료비"] < len(row) else 0.0
        l = _num(row[mapping["노무비"]]) if mapping.get("노무비") is not None and mapping["노무비"] < len(row) else 0.0
        e = _num(row[mapping["경비"]]) if mapping.get("경비") is not None and mapping["경비"] < len(row) else 0.0

        if m == 0 and l == 0 and e == 0:
            continue
        t.material += m
        t.labor += l
        t.expense += e
        t.line_count += 1

    if skipped:
        t.warnings.append(f"소계·합계로 보이는 {skipped}개 행을 이중계상 방지를 위해 제외했습니다.")
    if t.line_count == 0:
        t.warnings.append("금액이 잡힌 행이 없습니다. 열 매핑을 확인하세요.")
    if t.expense == 0 and t.labor > 0:
        t.warnings.append("경비가 0입니다. 기계경비·가설비가 별도 열에 있는지 확인하세요.")
    return t
